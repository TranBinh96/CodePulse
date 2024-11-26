import cv2
import json
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *


class ImageSelector(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Image Inspection Tool")

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setMouseTracking(True) # Enable mouse tracking for continuous feedback

        self.select_button = QPushButton("Select Image")
        self.select_button.clicked.connect(self.select_image)
        self.save_button = QPushButton("Save Configuration")
        self.save_button.clicked.connect(self.save_config)

        self.method_combo = QComboBox()
        self.method_combo.addItems(["Template Matching", "Color Detection", "YOLO"])
        self.method_combo.currentIndexChanged.connect(self.method_changed)

        self.master_region_group = QGroupBox("Master Region")
        self.master_file_edit = QLineEdit()
        self.save_path_edit = QLineEdit()
        self.min_match_percent_spin = QSpinBox()
        self.min_match_percent_spin.setRange(0, 100)
        self.min_match_percent_spin.setValue(50)
        self.procode_edit = QLineEdit()
        master_region_layout = QFormLayout()
        master_region_layout.addRow("ProCode:", self.procode_edit)
        master_region_layout.addRow("Master File:", self.master_file_edit)
        master_region_layout.addRow("Save Path:", self.save_path_edit)
        master_region_layout.addRow("Min Match Percent:", self.min_match_percent_spin)
        self.master_region_group.setLayout(master_region_layout)

        self.checkpoints_group = QGroupBox("Checkpoints")
        self.point_id_edit = QLineEdit()
        self.min_match_edit = QLineEdit()
        self.hsv_min_edit = QLineEdit()
        self.hsv_max_edit = QLineEdit()
        checkpoints_layout = QFormLayout()
        checkpoints_layout.addRow("Point ID:", self.point_id_edit)
        checkpoints_layout.addRow("Min Match Percent:", self.min_match_edit)
        checkpoints_layout.addRow("HSV Min (H,S,V):", self.hsv_min_edit)
        checkpoints_layout.addRow("HSV Max (H,S,V):", self.hsv_max_edit)
        self.checkpoints_group.setLayout(checkpoints_layout)
        self.checkpoints_group.setVisible(False)  # Initially hidden


        self.points = []

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        layout.addWidget(self.select_button)
        layout.addWidget(self.method_combo)
        layout.addWidget(self.master_region_group)
        layout.addWidget(self.checkpoints_group)
        layout.addWidget(self.save_button)
        self.setLayout(layout)

        self.current_image = None
        self.selected_points = [] # List of (x,y) tuples of selected points


    def select_image(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp);;All Files (*)", options=options)
        if file_name:
            self.load_image(file_name)


    def load_image(self, file_path):
        self.current_image = cv2.imread(file_path)
        if self.current_image is None:
            QMessageBox.critical(self, "Error", "Failed to load image.")
            return
        self.current_image = cv2.cvtColor(self.current_image, cv2.COLOR_BGR2RGB)
        h, w, ch = self.current_image.shape
        bytesPerLine = ch * w
        qImg = QImage(self.current_image.data, w, h, bytesPerLine, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qImg)
        self.image_label.setPixmap(pixmap)

    def method_changed(self, index):
        self.master_region_group.setVisible(index == 0)
        self.checkpoints_group.setVisible(index == 1)
        self.selected_points = []
        self.update_image()

    def save_config(self):
        filepath = r"D:\config.xlsx"
        try:
            self.save_excel_data(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving configuration: {e}")

    def save_excel_data(self, filepath):
        try:
            wb = load_workbook(filepath)
        except FileNotFoundError:
            wb = self.create_excel_file(filepath)
            if wb is None:
                return

        method = self.method_combo.currentText()
        if method == "Template Matching":
            self.save_master_region(wb)
        elif method == "Color Detection":
            self.save_checkpoint(wb)

        wb.save(filepath)

    def save_master_region(self, wb):
        ws = wb['MasterRegion']
        if len(self.selected_points) < 2:
            QMessageBox.warning(self, "Warning", "Please select at least two points for Master Region.")
            return

        # Prepare data in the desired order
        data = [
            self.procode_edit.text(),
            self.selected_points[0][0],
            self.selected_points[0][1],
            self.selected_points[-1][0],
            self.selected_points[-1][1],
            self.selected_points[0][0],
            self.selected_points[0][1],
            self.selected_points[-1][0],
            self.selected_points[-1][1],
            self.min_match_percent_spin.value(),
            self.master_file_edit.text(),
            self.save_path_edit.text()
        ]
        ws.append(data)

    def save_checkpoint(self, wb):
        ws = wb['Checkpoints']

        # Kiểm tra nếu có ít nhất một điểm được chọn
        if not self.selected_points:
            QMessageBox.warning(self, "Warning", "Please select at least one point for Checkpoint.")
            return

        # Lặp qua tất cả các điểm đã chọn để thêm dữ liệu thực tế
        for point in self.selected_points:
            bottom_right_y = point[1]  # Lấy tọa độ Y của điểm đã chọn
            hsv_range_min = self.hsv_min_edit.text()  # Giá trị từ trường nhập về màu HSV mini
            hsv_range_max = self.hsv_max_edit.text()  # Giá trị từ trường nhập về màu HSV max
            yolo_data = ""  # Giả sử không có dữ liệu YOLO
            min_match = self.min_match_edit.text()  # Giá trị nhập từ người dùng

            # Thêm dữ liệu vào Excel
            ws.append([bottom_right_y, hsv_range_min, hsv_range_max, yolo_data, min_match])

    def create_excel_file(self, filepath):
        try:
            wb = Workbook()

            # Xóa trang Sheet1 mặc định
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Tạo trang MasterRegion
            ws_mr = wb.create_sheet("MasterRegion", 0)  # Đảm bảo là trang đầu tiên
            # Thiết lập tiêu đề trong thứ tự yêu cầu
            ws_mr.append([
                "ProCode", "SearchTopLeftX", "SearchTopLeftY", "SearchBottomRightX",
                "SearchBottomRightY", "TopLeftX", "TopLeftY", "BottomRightX",
                "BottomRightY", "MinMatchPercent", "MasterFile", "SavePath"
            ])

            # Tạo trang Checkpoints
            ws_cp = wb.create_sheet("Checkpoints", 1)  # Trang thứ hai
            # Thiết lập tiêu đề mới trong thứ tự yêu cầu
            ws_cp.append([
                "BottomRightY", "HSVRangeMin", "HSVRangeMax", "YoloData", "MinMatch"
            ])

            # Lưu tập tin Excel
            wb.save(filepath)
            return wb
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error creating Excel file: {e}")
            return None

    def mousePressEvent(self, event):
        if self.current_image is not None:
            x = event.x() - self.image_label.x() # Adjust for label position
            y = event.y() - self.image_label.y() # Adjust for label position
            self.selected_points.append((x,y))
            self.update_image()


    def update_image(self):
        img_with_points = self.current_image.copy()
        for x, y in self.selected_points:
            cv2.circle(img_with_points, (x, y), 5, (0, 0, 255), -1)  # Draw red circles

        h, w, ch = img_with_points.shape
        bytesPerLine = ch * w
        qImg = QImage(img_with_points.data, w, h, bytesPerLine, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qImg)
        self.image_label.setPixmap(pixmap)



if __name__ == "__main__":
    app = QApplication([])
    window = ImageSelector()
    window.show()
    app.exec_()
